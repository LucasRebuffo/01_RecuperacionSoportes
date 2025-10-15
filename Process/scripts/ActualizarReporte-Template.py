import logging

log_path = "{gblRutaLogs}"

task = tmp_global_obj["profile"]["name"]  # type: ignore

logging.basicConfig(
    filename=log_path,
    level=logging.INFO,
    format="%(asctime)s | %(levelname)s | Task: {task} | Desc: %(message)s".format(
        task=task
    ),
    datefmt="%d-%m-%Y %H:%M:%S",
)
# ----------------------------------------------------------
from openpyxl import load_workbook

error = None

ruta_reporte = "{gblRutaReporte}"
logging.info(f"Se va a actualizar el archivo: {ruta_reporte}.")

# Abro el archivo de Excel
try:
    workbook = load_workbook(ruta_reporte)
    logging.info(f"Se carga el workbook.")

    sheet = workbook["{loc006Proveedor}".upper()]
    
    logging.info(f"Se carga la hoja: {sheet.title}.")

    estado = "{gblEstadoProcesamiento}"

    # Buscar numero de fila actual por valor de columna B
    for fila in range(1, sheet.max_row + 1):
        aux = sheet[f"B{fila}"].value.strip()
        aux2 = "{loc006NumeroComprobante}".strip()
        logging.info(f"Comparacion {aux} == {aux2} en fila {fila}.")
        if sheet[f"B{fila}"].value.strip() == "{loc006NumeroComprobante}".strip():  
            fila_actual = fila
            break
    else:
        logging.warning("No se encontró la fila con el valor especificado en la columna B.")
        logging.info(f"La ultima fila del excel es: {sheet.max_row}.")
        # Si no se encuentra, se usa la ultima fila
        fila_actual = sheet.max_row + 1

    logging.info(f"La fila actual es: {fila_actual}.")
 
    # Modifico el archivo (ejemplo: escribo un valor en la celda A1)
    sheet[f"A{fila_actual}"] = "{loc006Proveedor}"
    sheet[f"B{fila_actual}"] = "{loc006NumeroComprobante}"
    sheet[f"C{fila_actual}"] = "{loc006BaseImponible}"
    sheet[f"D{fila_actual}"] = "{loc006IVAAlicuota}"
    sheet[f"E{fila_actual}"] = "{loc006OtrosCargos}"
    sheet[f"F{fila_actual}"] = "{loc006TotalAPagar}"
    sheet[f"G{fila_actual}"] = "{loc006PeriodoFacturacion}"
    sheet[f"H{fila_actual}"] = "{loc006Medidor}"
    sheet[f"I{fila_actual}"] = "{loc006ResultadoExtraccionPDF}"
    sheet[f"J{fila_actual}"] = "{loc006ResultadoSAP}"
    sheet[f"K{fila_actual}"] = "{loc006ResultadoMoverAProcesados}"
    sheet[f"L{fila_actual}"] = "{loc006NumeroRegistroSAP}"
    sheet[f"M{fila_actual}"] = "{gblEstadoProcesamiento}"
    sheet[f"N{fila_actual}"] = "{loc006Observaciones}"
    sheet[f"O{fila_actual}"] = "{loc006FechaHoraInicio}"
    sheet[f"P{fila_actual}"] = "{loc006FechaHoraFin}"
    sheet[f"Q{fila_actual}"] = "{loc006TiempoProcesamiento}"
    
    # Guardo los cambios
    workbook.save(ruta_reporte)
    logging.info("El archivo de reporte ha sido actualizado correctamente.")

except FileNotFoundError as e:
    logging.warning(f"No se encontró el archivo en la ruta: {ruta_reporte}")
    error = e
except Exception as e:
    logging.warning(f"Error al actualizar el archivo: {str(e)}")
    error = e
finally:
    # Cierro el archivo
    workbook.close()

    # -----------------------------------------
    # Me aseguro de cerrar los handlers para evitar que se repita el log
    for handler in logging.root.handlers[:]:
        handler.close()
        logging.root.removeHandler(handler)

    if error:
        raise(error)
    