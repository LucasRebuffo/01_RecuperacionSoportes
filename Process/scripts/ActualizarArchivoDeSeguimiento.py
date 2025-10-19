global logger

from VortexLibrary import logger

logger.log_path = "{gblRutaLogs}"
logger.task_name = tmp_global_obj["profile"]["name"]

def actualizar_fila_especifica(input_data, archivo_excel):
    """
    Actualiza una fila especifica en el archivo de seguimiento basandose en la clave compuesta
    (tipo_docto + nro_docto + documento_cruce) y agrega una columna de fecha de procesamiento.
    
    Args:
        input_data (dict): Diccionario con los datos a actualizar
        archivo_excel (str): Ruta del archivo Excel a actualizar
    """
    import openpyxl
    from openpyxl import load_workbook
    import os
    from datetime import datetime

    try:
        # Verificar que el archivo existe
        if not os.path.exists(archivo_excel):
            logger.error(f"Error: El archivo '{archivo_excel}' no existe.")
            raise Exception(f"Error: El archivo '{archivo_excel}' no existe.")

        # Leer el archivo Excel existente
        workbook = load_workbook(archivo_excel)
        worksheet = workbook.active

        # Obtener la primera fila como encabezados
        headers = []
        for col in range(1, worksheet.max_column + 1):
            cell_value = worksheet.cell(row=1, column=col).value
            if cell_value is None:
                break
            headers.append(str(cell_value))
        

        normalized_headers = headers
        
        # Crear diccionario de mapeo de columnas
        column_mapping = {}
        for i, header in enumerate(normalized_headers):
            column_mapping[header] = i + 1  # +1 porque openpyxl usa indices basados en 1
        
        # Asegurar que todas las columnas necesarias existen
        columnas_requeridas = ['Empresa', 'Tipo Docto', 'Nro Docto', 'Detalle', 'Documento Cruce', 'Estado Procesamiento', 'Fecha Procesamiento','Observaciones']
        for col in columnas_requeridas:
            if col not in column_mapping:
                # Agregar nueva columna al final
                new_col = worksheet.max_column + 1
                worksheet.cell(row=1, column=new_col, value=col)
                column_mapping[col] = new_col

        # Contar filas con datos (excluyendo encabezados)
        data_rows = 0
        for row in range(2, worksheet.max_row + 1):
            empresa_value = worksheet.cell(row=row, column=column_mapping.get('Empresa', 1)).value
            if empresa_value is not None and str(empresa_value).strip() != '':
                data_rows += 1

        logger.info(f"Archivo cargado: {data_rows} filas de datos")
        logger.info(f"Columnas disponibles: {list(column_mapping.keys())}")

        # Crear clave compuesta para identificar la fila a actualizar
        clave_buscar = str(input_data['Tipo Docto']) + '_' + str(input_data['Nro Docto']) + '_' + str(input_data['Documento Cruce'])
        
        # Buscar todas las filas que coinciden con la clave
        filas_encontradas = []
        claves_encontradas = []
        
        for row in range(2, worksheet.max_row + 1):
            tipo_docto_value = worksheet.cell(row=row, column=column_mapping.get('Tipo Docto', 1)).value
            nro_docto_value = worksheet.cell(row=row, column=column_mapping.get('Nro Docto', 1)).value
            documento_cruce_value = worksheet.cell(row=row, column=column_mapping.get('Documento Cruce', 1)).value
            
            if tipo_docto_value is not None and nro_docto_value is not None and documento_cruce_value is not None:
                clave_actual = str(tipo_docto_value) + '_' + str(nro_docto_value) + '_' + str(documento_cruce_value)
                claves_encontradas.append(clave_actual)
                
                if clave_actual == clave_buscar:
                    filas_encontradas.append(row)
        
        logger.info(f"Filas encontradas: {filas_encontradas}")
        
        if len(filas_encontradas) == 0:
            logger.warning(f"No se encontro ninguna fila con la clave: {clave_buscar}")
            logger.info("Claves disponibles en el archivo:")
            for idx, clave in enumerate(claves_encontradas[:10]):  # Mostrar solo las primeras 10
                logger.info(f"  {idx+1}. {clave}")
            return False
        
        logger.info(f"Se encontraron {len(filas_encontradas)} filas con la clave compuesta: {clave_buscar}")
        logger.info(f"Clave compuesta: Tipo Docto + Nro Docto + Documento Cruce")
        logger.info(f"Indices de filas encontradas: {filas_encontradas}")
        
        # Mostrar datos actuales de todas las filas encontradas
        for i, fila_indice in enumerate(filas_encontradas):
            logger.info(f"Datos actuales de la fila {i+1} (indice {fila_indice}):")
            for col in columnas_requeridas:
                col_index = column_mapping.get(col, 1)
                cell_value = worksheet.cell(row=fila_indice, column=col_index).value
                logger.info(f"  {col}: {cell_value}")

        # Actualizar todas las filas encontradas con los nuevos datos
        filas_actualizadas = 0
        for fila_indice in filas_encontradas:
            logger.info(f"Actualizando fila {fila_indice}...")
            
            # Actualizar la fila con los nuevos datos
            for columna, valor in input_data.items():
                if columna in column_mapping:
                    col_index = column_mapping[columna]
                    worksheet.cell(row=fila_indice, column=col_index, value=valor)
                    logger.info(f"  Actualizado {columna}: {valor}")

            # Agregar columna de fecha de procesamiento si no existe
            if 'Fecha Procesamiento' not in column_mapping:
                new_col = worksheet.max_column + 1
                worksheet.cell(row=1, column=new_col, value='Fecha Procesamiento')
                column_mapping['Fecha Procesamiento'] = new_col

            # Actualizar la fecha de procesamiento
            fecha_actual = datetime.now().strftime("%d-%m-%Y %H:%M:%S")
            fecha_col_index = column_mapping['Fecha Procesamiento']
            worksheet.cell(row=fila_indice, column=fecha_col_index, value=fecha_actual)
            
            filas_actualizadas += 1
            logger.info(f"Fila {fila_indice} actualizada exitosamente")

        # Guardar el archivo actualizado
        workbook.save(archivo_excel)

        logger.info(f"Archivo actualizado exitosamente: {archivo_excel}")
        logger.info(f"Fecha de procesamiento: {fecha_actual}")
        logger.info(f"Total de filas actualizadas: {filas_actualizadas}")
        logger.info(f"Total de filas en el archivo: {worksheet.max_row}")


    except Exception as e:
        logger.error(f"Error al actualizar la fila: {str(e)}")
        raise e

# Datos de entrada del ejemplo
input_data = {gblItemActual}

# Ejecutar la actualizacion
resultado = actualizar_fila_especifica(input_data, "{gblRutaInput}"+"/Archivo de Seguimiento.xlsx")

logger.info("Actualizacion completada exitosamente")
