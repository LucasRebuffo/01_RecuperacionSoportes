import pandas as pd
import os
from openpyxl import load_workbook, Workbook

def actualizar_archivo_seguimiento(input_data, archivo_excel="Archivo de Seguimiento.xlsx"):
    """
    Actualiza el archivo de seguimiento con los datos del input.
    Si ya existe una fila con la misma clave compuesta (tipo_docto + nro_docto),
    no actualiza estado_procesamiento ni documento_cruce.
    """
    
    try:
        # Intentar leer el archivo Excel existente
        df_existente = None
        
        if os.path.exists(archivo_excel):
            try:
                # Leer el archivo con openpyxl para entender mejor la estructura
                wb = load_workbook(archivo_excel, data_only=True)
                ws = wb.active
                
                # Buscar la fila de encabezados (que contenga 'Empresa')
                header_row = None
                for row_idx in range(1, ws.max_row + 1):
                    if ws.cell(row=row_idx, column=1).value == 'Empresa':
                        header_row = row_idx
                        break
                
                if header_row:
                    # Leer datos desde la fila de encabezados
                    data_rows = []
                    for row_idx in range(header_row + 1, ws.max_row + 1):
                        row_data = []
                        for col_idx in range(1, 7):  # 6 columnas
                            cell_value = ws.cell(row=row_idx, column=col_idx).value
                            if cell_value is None:
                                break
                            row_data.append(cell_value)
                        
                        if len(row_data) == 6 and any(row_data):  # Solo filas con datos completos
                            data_rows.append(row_data)
                    
                    if data_rows:
                        df_existente = pd.DataFrame(data_rows, columns=[
                            'empresa', 'tipo_docto', 'nro_docto', 'detalle', 
                            'documento_cruce', 'estado_procesamiento'
                        ])
                        print(f"Archivo leído exitosamente: {len(df_existente)} filas de datos")
                    else:
                        print("Archivo existe pero no tiene datos válidos")
                        df_existente = pd.DataFrame(columns=[
                            'empresa', 'tipo_docto', 'nro_docto', 'detalle', 
                            'documento_cruce', 'estado_procesamiento'
                        ])
                else:
                    print("No se encontró fila de encabezados válida")
                    df_existente = pd.DataFrame(columns=[
                        'empresa', 'tipo_docto', 'nro_docto', 'detalle', 
                        'documento_cruce', 'estado_procesamiento'
                    ])
                    
            except Exception as e:
                print(f"Error al leer archivo existente: {e}")
                df_existente = pd.DataFrame(columns=[
                    'empresa', 'tipo_docto', 'nro_docto', 'detalle', 
                    'documento_cruce', 'estado_procesamiento'
                ])
        else:
            print("Archivo no existe, se creará uno nuevo")
            df_existente = pd.DataFrame(columns=[
                'empresa', 'tipo_docto', 'nro_docto', 'detalle', 
                'documento_cruce', 'estado_procesamiento'
            ])
        
        # Convertir NaN a string vacío
        df_existente = df_existente.fillna('')
        
        print(f"Archivo existente cargado: {len(df_existente)} filas")
        print(f"Columnas: {list(df_existente.columns)}")
        
        # Convertir input_data a DataFrame
        df_nuevo = pd.DataFrame(input_data)
        
        print(f"Datos de entrada: {len(df_nuevo)} filas")
        print(f"Columnas entrada: {list(df_nuevo.columns)}")
        
        # Crear clave compuesta para identificar filas existentes
        df_existente['clave_compuesta'] = df_existente['tipo_docto'].astype(str) + '_' + df_existente['nro_docto'].astype(str)
        df_nuevo['clave_compuesta'] = df_nuevo['tipo_docto'].astype(str) + '_' + df_nuevo['nro_docto'].astype(str)
        
        # Filas a agregar (nuevas)
        filas_nuevas = df_nuevo[~df_nuevo['clave_compuesta'].isin(df_existente['clave_compuesta'])]
        
        # Filas a actualizar (existentes) - solo actualizar empresa y detalle
        filas_actualizar = df_nuevo[df_nuevo['clave_compuesta'].isin(df_existente['clave_compuesta'])]
        
        print(f"Filas nuevas a agregar: {len(filas_nuevas)}")
        print(f"Filas existentes a actualizar: {len(filas_actualizar)}")
        
        # Actualizar filas existentes (solo empresa y detalle)
        for _, fila_nueva in filas_actualizar.iterrows():
            clave = fila_nueva['clave_compuesta']
            mask = df_existente['clave_compuesta'] == clave
            
            # Actualizar solo empresa y detalle, mantener estado_procesamiento y documento_cruce
            df_existente.loc[mask, 'empresa'] = fila_nueva['empresa']
            df_existente.loc[mask, 'detalle'] = fila_nueva['detalle']
            print(f"Actualizada fila existente: {clave}")
        
        # Agregar filas nuevas
        if len(filas_nuevas) > 0:
            # Eliminar la columna clave_compuesta antes de concatenar
            filas_nuevas_clean = filas_nuevas.drop('clave_compuesta', axis=1)
            df_existente = pd.concat([df_existente, filas_nuevas_clean], ignore_index=True)
            print(f"Agregadas {len(filas_nuevas)} filas nuevas")
        
        # Eliminar la columna clave_compuesta del resultado final
        df_existente = df_existente.drop('clave_compuesta', axis=1)
        
        # Guardar el archivo actualizado
        wb = Workbook()
        ws = wb.active
        
        # Escribir encabezados en la fila 4
        headers = ['Empresa', 'Tipo\nDocto', 'Nro\nDocto', 'Detalle', 'DOCUMENTO CRUCE', 'Estado Procesamiento']
        for col, header in enumerate(headers, 1):
            ws.cell(row=4, column=col).value = header
        
        # Escribir datos
        for row_idx, (_, row) in enumerate(df_existente.iterrows(), 5):
            ws.cell(row=row_idx, column=1).value = row['empresa']
            ws.cell(row=row_idx, column=2).value = row['tipo_docto']
            ws.cell(row=row_idx, column=3).value = row['nro_docto']
            ws.cell(row=row_idx, column=4).value = row['detalle']
            ws.cell(row=row_idx, column=5).value = row['documento_cruce']
            ws.cell(row=row_idx, column=6).value = row['estado_procesamiento']
        
        # Guardar el archivo
        wb.save(archivo_excel)
        
        print(f"Archivo actualizado exitosamente: {archivo_excel}")
        print(f"Total de filas en el archivo: {len(df_existente)}")
        
        return df_existente.to_dict('records')
        
    except Exception as e:
        print(f"Error al actualizar el archivo: {str(e)}")
        return None

# Datos de entrada del ejemplo original
input_data = [
    {'empresa': 'Ticenergi', 'tipo_docto': 'NQ', 'nro_docto': 26, 'detalle': 'COMPROBANTE DE CAUSACION', 'documento_cruce': '', 'estado_procesamiento': 'Pendiente'}, 
    {'empresa': 'Ticenergi', 'tipo_docto': 'NQ', 'nro_docto': 27, 'detalle': 'COMPROBANTE DE CAUSACION', 'documento_cruce': '', 'estado_procesamiento': 'Pendiente'}
]

# Ejecutar la actualización
resultado = actualizar_archivo_seguimiento(input_data)

if resultado:
    print("\n=== CONTENIDO ACTUALIZADO DEL ARCHIVO ===")
    for i, fila in enumerate(resultado, 1):
        print(f"Fila {i}: {fila}")
else:
    print("No se pudo actualizar el archivo.")
