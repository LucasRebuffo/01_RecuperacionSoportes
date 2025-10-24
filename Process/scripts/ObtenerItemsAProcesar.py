def actualizar_archivo_seguimiento(input_data, archivo_excel, template_excel):
    """
    Actualiza el archivo de seguimiento con los datos del input.
    Si ya existe una fila con la misma clave compuesta (tipo_docto + nro_docto),
    no hace nada (NO actualiza, NO sobreescribe absolutamente ninguna columna de ninguna fila existente).
    Solo inserta los registros nuevos.
    Si el archivo no existe lo crea a partir de un template que tiene el estilo de tabla correcto.
    """

    import os
    import shutil
    import openpyxl
    from openpyxl.utils import get_column_letter
    from VortexLibrary import logger

    logger.log_path = "{gblRutaLogs}"
    logger.task_name = tmp_global_obj["profile"]["name"]

    try:
        # Si el archivo de seguimiento NO existe, crearlo copiando el template con formato/tablas
        if not os.path.exists(archivo_excel):
            logger.info(
                "El archivo de seguimiento no existe, se creará a partir del template: "
                + str(template_excel)
            )
            if not os.path.exists(template_excel):
                logger.error(
                    f"El archivo template '{template_excel}' no existe. No se puede crear el archivo de seguimiento."
                )
                SetVar("gblItemsAProcesar", [])
                return
            shutil.copyfile(template_excel, archivo_excel)
            logger.info(
                f"Archivo de seguimiento creado a partir del template: {archivo_excel}"
            )

        # Leer el archivo Excel existente usando openpyxl
        try:
            wb = openpyxl.load_workbook(archivo_excel)
            ws = wb.active

            # Encontrar la fila de encabezado
            header_row_idx = None
            expected_first_col = "Empresa"

            for row in ws.iter_rows(min_row=1, max_col=1):
                if row[0].value == expected_first_col:
                    header_row_idx = row[0].row
                    break

            if not header_row_idx:
                header_row_idx = 1  # Fallback: primera fila

            # Obtener encabezados
            header_row = [cell.value for cell in ws[header_row_idx]]
            logger.info(f"Encabezados encontrados: {header_row}")

            # Leer datos existentes
            datos_existentes = []
            for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
                if (
                    row and row[0] is not None
                ):  # Solo filas con datos válidos (empresa no vacía)
                    datos_existentes.append(list(row))

            logger.info(
                f"Archivo leído exitosamente: {len(datos_existentes)} filas de datos"
            )

            # Asegurar que todas las columnas necesarias existen
            columnas_requeridas = [
                "Empresa",
                "Tipo Docto",
                "Nro Docto",
                "Detalle",
                "Documento Cruce",
                "Nro de Proceso",
                "Doc Alt",
                "Estado Procesamiento",
                "Fecha Procesamiento",
                "Observaciones",
            ]
            for col in columnas_requeridas:
                if col not in header_row:
                    header_row.append(col)
                    # Agregar columna vacía a datos existentes
                    for fila in datos_existentes:
                        fila.append("")

            logger.info(f"Archivo existente cargado: {len(datos_existentes)} filas")
            logger.info(f"Columnas: {header_row}")

            # Convertir input_data a lista de listas
            datos_nuevos = []
            for item in input_data:
                fila = []
                for col in header_row:
                    if col == "Empresa":
                        fila.append(item.get("empresa", ""))
                    elif col == "Tipo Docto":
                        fila.append(item.get("tipo_docto", ""))
                    elif col == "Nro Docto":
                        fila.append(item.get("nro_docto", ""))
                    elif col == "Detalle":
                        fila.append(item.get("detalle", ""))
                    elif col == "Documento Cruce":
                        fila.append(item.get("documento_cruce", ""))
                    elif col == "Nro de Proceso":
                        fila.append(item.get("nro_proceso", ""))
                    elif col == "Doc Alt":
                        fila.append(item.get("doc_alt", ""))
                    elif col == "Estado Procesamiento":
                        fila.append(item.get("estado_procesamiento", ""))
                    elif col == "Fecha Procesamiento":
                        fila.append(item.get("fecha_procesamiento", ""))
                    elif col == "Observaciones":
                        fila.append(item.get("observaciones", ""))
                    else:
                        fila.append("")
                datos_nuevos.append(fila)

            logger.info(f"Datos de entrada: {len(datos_nuevos)} filas")

            # Crear claves compuestas para identificar filas existentes
            claves_existentes = set()
            for fila in datos_existentes:
                if (
                    len(fila) >= 3
                ):  # Asegurar que hay al menos empresa, tipo_docto, nro_docto
                    tipo_idx = (
                        header_row.index("Tipo Docto")
                        if "Tipo Docto" in header_row
                        else 1
                    )
                    nro_idx = (
                        header_row.index("Nro Docto")
                        if "Nro Docto" in header_row
                        else 2
                    )
                    clave = str(fila[tipo_idx]) + "_" + str(fila[nro_idx])
                    claves_existentes.add(clave)

            # Filtrar datos nuevos (que no existen)
            datos_a_agregar = []
            for fila in datos_nuevos:
                if len(fila) >= 3:
                    tipo_idx = (
                        header_row.index("Tipo Docto")
                        if "Tipo Docto" in header_row
                        else 1
                    )
                    nro_idx = (
                        header_row.index("Nro Docto")
                        if "Nro Docto" in header_row
                        else 2
                    )
                    clave = str(fila[tipo_idx]) + "_" + str(fila[nro_idx])
                    if clave not in claves_existentes:
                        datos_a_agregar.append(fila)

            logger.info(f"Filas nuevas a agregar: {len(datos_a_agregar)}")

            # Agregar filas nuevas al archivo
            if len(datos_a_agregar) > 0:
                # Detectar si existe una tabla en el archivo
                tabla_existente = None
                for table in ws.tables.values():
                    # Parsear el rango de la tabla para obtener filas de inicio y fin
                    try:
                        # El formato del rango es "A1:C10" por ejemplo
                        rango_parts = table.ref.split(":")
                        if len(rango_parts) == 2:
                            # Extraer número de fila del inicio y fin del rango
                            inicio_fila = int(
                                "".join(filter(str.isdigit, rango_parts[0]))
                            )
                            fin_fila = int("".join(filter(str.isdigit, rango_parts[1])))

                            # Verificar si la tabla incluye la fila de encabezado
                            if inicio_fila <= header_row_idx <= fin_fila:
                                tabla_existente = table
                                break
                    except (ValueError, IndexError) as e:
                        logger.warning(
                            f"Error al parsear rango de tabla {table.name}: {e}"
                        )
                        continue

                # Log: inicio del borrado de filas de datos bajo el header
                logger.info(f"Eliminando filas de datos debajo del header en la fila indice {header_row_idx}")
                max_row = ws.max_row
                data_start_idx = header_row_idx + 1
                logger.info(f"max_row en hoja: {max_row}, data_start_idx: {data_start_idx}")
                if max_row >= data_start_idx:
                    logger.info(f"Eliminando filas desde {data_start_idx} hasta {max_row}")
                    ws.delete_rows(data_start_idx, max_row - data_start_idx + 1)

                # Concatenando datos existentes y nuevos
                logger.info("Concatenando filas existentes y nuevas para escribir")
                todos_los_datos = datos_existentes + datos_a_agregar
                logger.info(f"Cantidad total de filas para escribir (antes de quitar duplicados): {len(todos_los_datos)}")

                # Eliminar duplicados basados en Tipo Docto, Nro Docto, DOCUMENTO CRUCE
                col_tipo = (
                    header_row.index("Tipo Docto")
                    if "Tipo Docto" in header_row
                    else None
                )
                col_nro = (
                    header_row.index("Nro Docto") if "Nro Docto" in header_row else None
                )
                col_cruce = (
                    header_row.index("Documento Cruce")
                    if "Documento Cruce" in header_row
                    else None
                )
                logger.info(f"Indices para deduplicar: {col_tipo}, {col_nro}, {col_cruce}")
                if None not in (col_tipo, col_nro, col_cruce):
                    seen_keys = set()
                    datos_sin_duplicados = []
                    dedup_total = 0
                    for fila in todos_los_datos:
                        if len(fila) > max(col_tipo, col_nro, col_cruce):
                            clave = (
                                str(fila[col_tipo]).strip()
                                + "_"
                                + str(fila[col_nro]).strip()
                                + "_"
                                + str(fila[col_cruce]).strip()
                            )
                            if clave not in seen_keys:
                                seen_keys.add(clave)
                                datos_sin_duplicados.append(fila)
                            else:
                                dedup_total += 1
                                logger.info(f"Fila duplicada eliminada por clave: {clave}")
                    if dedup_total > 0:
                        logger.info(f"Cantidad total de filas duplicadas eliminadas: {dedup_total}")
                    else:
                        logger.info("No se encontraron filas duplicadas")
                    todos_los_datos = datos_sin_duplicados
                else:
                    logger.warning("No se encontraron todas las columnas necesarias para eliminar duplicados. No se realizará eliminacion.")

                logger.info(f"Escribiendo {len(todos_los_datos)} filas en hoja, comenzando en la fila {data_start_idx}")
                for r_idx, fila_datos in enumerate(
                    todos_los_datos, start=data_start_idx
                ):
                    logger.info(f"Escribiendo fila {r_idx} en hoja")
                    for c_idx, valor in enumerate(fila_datos, start=1):
                        logger.info(f"  Valor fila {r_idx} columna {c_idx}: {valor}")
                        ws.cell(row=r_idx, column=c_idx, value=valor)

                # Si existe una tabla, actualizar su rango para incluir los nuevos datos
                if tabla_existente:
                    try:
                        # Calcular el nuevo rango de la tabla
                        num_columnas = len(header_row)
                        nueva_fila_final = header_row_idx + len(todos_los_datos)
                        nuevo_rango = f"{get_column_letter(1)}{header_row_idx}:{get_column_letter(num_columnas)}{nueva_fila_final}"

                        # Actualizar el rango de la tabla
                        tabla_existente.ref = nuevo_rango
                        logger.info(f"Tabla actualizada con nuevo rango: {nuevo_rango}")

                    except Exception as e:
                        logger.warning(
                            f"No se pudo actualizar el rango de la tabla: {e}"
                        )

                wb.save(archivo_excel)
                logger.info(f"Agregadas {len(datos_a_agregar)} filas nuevas")
            else:
                logger.info("No hay filas nuevas por agregar.")

            logger.info(f"Archivo actualizado exitosamente: {archivo_excel}")
            logger.info(
                f"Total de filas en el archivo: {len(datos_existentes) + len(datos_a_agregar)}"
            )

            # Preparar datos para SetVar
            todos_los_datos = datos_existentes + datos_a_agregar

            # Eliminar duplicados basados en Tipo Docto, Nro Docto, Documento Cruce
            logger.info("Eliminando duplicados basados en 'Tipo Docto', 'Nro Docto', 'Documento Cruce'...")
            # Buscar los índices de las columnas
            col_tipo = (
                header_row.index("Tipo Docto") if "Tipo Docto" in header_row else None
            )
            col_nro = (
                header_row.index("Nro Docto") if "Nro Docto" in header_row else None
            )
            col_cruce = (
                header_row.index("Documento Cruce")
                if "Documento Cruce" in header_row
                else None
            )
            logger.info(f"Indices columnas - Tipo Docto: {col_tipo}, Nro Docto: {col_nro}, Documento Cruce: {col_cruce}")
            if None not in (col_tipo, col_nro, col_cruce):
                seen_keys = set()
                datos_sin_duplicados = []
                cant_duplicados = 0
                for fila in todos_los_datos:
                    if len(fila) > max(col_tipo, col_nro, col_cruce):
                        clave = (
                            str(fila[col_tipo]).strip()
                            + "_"
                            + str(fila[col_nro]).strip()
                            + "_"
                            + str(fila[col_cruce]).strip()
                        )
                        if clave not in seen_keys:
                            seen_keys.add(clave)
                            datos_sin_duplicados.append(fila)
                        else:
                            cant_duplicados += 1
                            logger.info(f"Fila duplicada encontrada con clave {clave}, se elimina de los datos a procesar.")
                logger.info(f"Total de filas duplicadas eliminadas: {cant_duplicados}")
                logger.info(f"Total de filas sin duplicados: {len(datos_sin_duplicados)}")
                todos_los_datos = datos_sin_duplicados
            else:
                logger.warning("No se encontraron todas las columnas necesarias para eliminar duplicados. No se realizará eliminación.")

            logger.info(f"Preparando datos finales para SetVar. Total de filas: {len(todos_los_datos)}")
            datos_para_setvar = []
            for fila in todos_los_datos:
                dict_fila = {}
                for i, col in enumerate(header_row):
                    if i < len(fila):
                        dict_fila[col] = fila[i]
                    else:
                        dict_fila[col] = ""
                datos_para_setvar.append(dict_fila)
            logger.info(f"Número de registros preparados para SetVar: {len(datos_para_setvar)}")

            SetVar("gblItemsAProcesar", datos_para_setvar)

            # Cerrar el archivo explícitamente
            wb.close()

        except Exception as e:
            # Asegurar que el archivo se cierre en caso de error
            try:
                if "wb" in locals():
                    wb.close()
            except:
                pass
            logger.error(f"Error al leer archivo existente: {e}")
            raise e

    except Exception as e:
        logger.error(f"Error al actualizar el archivo: {str(e)}")
        raise e


# Datos de entrada del ejemplo original
input_data = {gblDatosArchivoBase}
template_excel = "{gblRutaTemplates}/Template Archivo de Seguimiento.xlsx"

# Ejecutar la actualización
actualizar_archivo_seguimiento(
    input_data, "{gblRutaInput}" + "/Archivo de Seguimiento.xlsx", template_excel
)
